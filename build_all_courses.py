"""Curriculum Instructor Agent — reads Academy spreadsheet, generates course HTML pages.
Pattern: Research (spreadsheet) → Index (AI Results sheet) → Generate (HTML pages)
"""
import openpyxl, os, html as h

SRC = '/Users/patmini/Downloads/Academy spreadsheet.xlsx'
OUT = '/tmp/artispreneur-landing/courses'

wb = openpyxl.load_workbook(SRC, data_only=True)

# Layer 2: MASTER INDEX — read AI Results sheet
ai_sheet = wb['All Courses - AI Results']
courses_index = {}
for row in ai_sheet.iter_rows(min_row=2, values_only=True):
    course_name, module_id, chapter, module_title = row[0], row[1], row[2], row[3]
    if not course_name: continue
    course_name = str(course_name).strip()
    if course_name not in courses_index:
        courses_index[course_name] = {'modules': [], 'category': 'Music Business'}
    courses_index[course_name]['modules'].append({
        'id': str(module_id),
        'chapter': str(chapter) if chapter else '',
        'title': str(module_title) if module_title else ''
    })

# Layer 3: GENERATE — build HTML pages from individual course sheets + index
os.makedirs(OUT, exist_ok=True)
generated = []

# Map sheet names to course titles
sheet_to_course = {}
for sname in wb.sheetnames:
    if sname in ['All Courses', 'All Courses - AI Results']: continue
    # Try to match with course index
    for cname in courses_index:
        clean_sheet = sname.lower().replace(' ','').replace('-','').replace('.','')
        clean_course = cname.lower().replace(' ','').replace('-','').replace('.','')
        if clean_course[:20] in clean_sheet or clean_sheet[:20] in clean_course:
            sheet_to_course[sname] = cname
            break

category_map = {
    'Brand': 'Branding', 'Collaborative': 'Streaming', 'PRO': 'Royalties',
    'Book': 'Marketing', 'Copyright': 'Legal', 'Catalogue': 'Business',
    'Catalog': 'Business', 'Distribute': 'Distribution', 'Tax': 'Finance',
    'Bank': 'Finance', 'LLC': 'Legal', 'CRM': 'Business',
    'Submit': 'Marketing', 'License': 'Licensing', 'E.I.N': 'Legal',
    'Register with': 'Royalties', 'Trademark': 'Legal', 'Network': 'Business'
}

slug_map = {
    'Brand Yourself': 'brand-yourself-as-artist',
    'Collaborative Playlist': 'add-music-collaborative-playlist',
    'Add Songs to your P.R.O.': 'add-songs-to-pro',
    'Book and Promote': 'promote-your-shows',
    'Copyright Your Music': 'copyright-your-music',
    'Create a Music Catalogue': 'create-music-catalogue',
    'Distribute Your Music': 'distribute-your-music',
    'File Business Taxes': 'file-business-taxes',
    'Set up Business Bank': 'setup-business-bank-account',
    'Set Up LLC': 'setup-llc',
    'Set up a CRM': 'setup-crm',
    'Submit Music to Blogs': 'submit-music-blogs-playlists',
    'License Your Music': 'license-your-music',
    'Register an E.I.N': 'register-ein',
    'Register with a P.R.O': 'register-with-pro',
    'Network in the Music': 'network-music-industry',
    'Trademark Your Music': 'trademark-your-music',
}

CSS = '''<style>
:root{--black:#09090b;--surface:#18181b;--card:#1c1c1f;--border:#27272a;--text:#fafafa;--text-muted:#a1a1aa;--text-dim:#71717a;--gold:#c9a227;--gold-light:#e8c84a;--gold-muted:rgba(201,162,39,0.12);--max-w:800px}
*{margin:0;padding:0;box-sizing:border-box}
html{-webkit-font-smoothing:antialiased}
body{font-family:'Inter',-apple-system,sans-serif;background:var(--black);color:var(--text);line-height:1.7}
a{color:var(--gold);text-decoration:none;transition:color .2s}
a:hover{color:var(--gold-light)}
.container{max-width:var(--max-w);margin:0 auto;padding:0 24px}
.navbar{position:fixed;top:0;left:0;right:0;z-index:1000;padding:0 24px;background:rgba(9,9,11,.92);backdrop-filter:blur(20px);border-bottom:1px solid var(--border)}
.navbar-inner{max-width:1100px;margin:0 auto;display:flex;align-items:center;justify-content:space-between;height:64px}
.nav-links{display:flex;align-items:center;gap:24px}
.nav-links a{color:var(--text-muted);font-size:13px;font-weight:500;white-space:nowrap}
.nav-links a:hover{color:var(--text)}
main{padding-top:100px;padding-bottom:80px}
.breadcrumb{font-size:13px;color:var(--text-dim);margin-bottom:24px}
.breadcrumb a{color:var(--text-dim)}
.breadcrumb a:hover{color:var(--gold)}
.course-title{font-family:'Playfair Display',serif;font-size:clamp(28px,3.5vw,42px);font-weight:900;margin-bottom:8px}
.course-meta{display:flex;gap:16px;margin-bottom:8px;font-size:13px}
.course-cat{padding:3px 10px;border-radius:100px;font-size:11px;font-weight:600;background:var(--gold-muted);color:var(--gold)}
.course-lessons{color:var(--text-dim)}
.course-desc{font-size:15px;color:var(--text-muted);margin-bottom:40px;line-height:1.7}
.module{background:var(--surface);border:1px solid var(--border);border-radius:12px;padding:28px;margin-bottom:20px}
.module-header{display:flex;align-items:center;gap:14px;margin-bottom:6px}
.module-num{font-size:11px;font-weight:700;color:var(--gold);background:var(--gold-muted);padding:3px 10px;border-radius:4px;white-space:nowrap}
.module-header h3{font-size:17px;font-weight:700;flex:1}
.module-chapter{font-size:12px;color:var(--text-dim);margin-bottom:14px;text-transform:uppercase;letter-spacing:.5px}
.module-content{font-size:14px;color:var(--text-muted);line-height:1.8;white-space:pre-wrap}
.cta{text-align:center;margin-top:48px;padding:36px;background:var(--surface);border:1px solid var(--border);border-radius:16px}
.cta h2{font-size:20px;font-weight:700;margin-bottom:8px}
.cta p{font-size:14px;color:var(--text-muted);margin-bottom:16px}
.cta .btn{display:inline-block;background:var(--gold);color:var(--black);padding:12px 28px;border-radius:8px;font-weight:700;font-size:14px}
.cta .btn:hover{background:var(--gold-light)}
footer{background:var(--card);border-top:1px solid var(--border);padding:32px 0;text-align:center;font-size:13px;color:var(--text-dim)}
@media(max-width:640px){.module-header{flex-wrap:wrap}.nav-links{display:none}}
</style>'''

for sname, cname in sheet_to_course.items():
    if sname not in wb.sheetnames: continue
    ws = wb[sname]
    
    # Find category
    cat = 'Music Business'
    for key, val in category_map.items():
        if key.lower() in cname.lower():
            cat = val
            break
    
    # Find slug
    slug = None
    for key, val in slug_map.items():
        if key.lower() in cname.lower():
            slug = val
            break
    
    if not slug:
        slug = cname.lower().replace(' ','-').replace(':','').replace('.','')[:50]
    
    # Read modules from individual sheet
    modules = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        mid = str(row[0]) if row[0] else ''
        chapter = str(row[1]) if row[1] else ''
        mtitle = str(row[2]) if row[2] else ''
        content = str(row[3]) if row[3] else ''
        if mid and mtitle:
            modules.append({'id': mid, 'chapter': chapter, 'title': mtitle, 'content': content})
    
    if not modules: continue
    
    # Build module HTML
    mod_html = ''
    for m in modules:
        mid_parts = m['id'].split('.')
        num = mid_parts[0] if mid_parts else m['id']
        clean_content = h.escape(m['content'])
        mod_html += f'''
    <div class="module">
      <div class="module-header">
        <span class="module-num">Module {m['id']}</span>
        <h3>{h.escape(m['title'])}</h3>
      </div>
      <p class="module-chapter">{h.escape(m['chapter'])}</p>
      <div class="module-content">{clean_content}</div>
    </div>'''
    
    html = f'''<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>{h.escape(cname)} — Artispreneur Academy</title>
<meta name="description" content="Learn {h.escape(cname.lower())} in this Artispreneur Academy course. {len(modules)} detailed modules.">
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800;900&family=Playfair+Display:wght@700;800;900&display=swap" rel="stylesheet">
{CSS}
</head>
<body>
<nav class="navbar">
  <div class="navbar-inner">
    <a href="../index.html"><img src="../logo.png" alt="Artispreneur" style="height:24px;display:block"></a>
    <div class="nav-links">
      <a href="../index.html#agents">Agents</a>
      <a href="../academy.html">Academy</a>
      <a href="../directory.html">Directory</a>
      <a href="../pricing.html">Pricing</a>
    </div>
  </div>
</nav>
<main class="container">
  <p class="breadcrumb"><a href="../index.html">Home</a> / <a href="../academy.html">Academy</a> / {h.escape(cname)}</p>
  <h1 class="course-title">{h.escape(cname)}</h1>
  <div class="course-meta">
    <span class="course-cat">{cat}</span>
    <span class="course-lessons">{len(modules)} lessons</span>
  </div>
  {mod_html}
  <div class="cta">
    <h2>Want more?</h2>
    <p>Get the full experience with video lessons on the Artispreneur Academy.</p>
    <a href="https://academy-build.vercel.app/#courses" class="btn">Open Full Academy →</a>
  </div>
</main>
<footer><p>&copy; 2026 Artispreneur · Art Means Business</p></footer>
</body>
</html>'''
    
    fpath = os.path.join(OUT, f"{slug}.html")
    with open(fpath, 'w') as f:
        f.write(html)
    generated.append((cname, slug, len(modules), cat))

for name, slug, count, cat in generated:
    print(f"  ✓ {cat:15s} | {count:2d} modules | {slug}.html")

print(f"\nGenerated {len(generated)} course pages from spreadsheet")
